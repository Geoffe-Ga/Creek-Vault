"""Spreadsheet ingestor for the Creek pipeline.

Implements §57 of the Creek Ontology — ingest XLSX workbooks (one
fragment per sheet) and CSV files (one fragment per file). The
spreadsheet backend is decoupled through the
:class:`SpreadsheetBackend` Protocol so callers can plug in any
implementation; tests inject a deterministic stub. The default
:class:`OpenpyxlBackend` lazily imports ``openpyxl`` so the rest of
the package — and the unit tests — run cleanly even when that
optional dependency is not installed. CSV parsing uses Python's
stdlib ``csv`` module and works without any optional dependency.

Sheets larger than :data:`SUMMARY_THRESHOLD` rows render as a
truncated table with the first :data:`SUMMARY_HEAD_ROWS` and last
:data:`SUMMARY_TAIL_ROWS` rows plus the total row count, so giant
exports stay legible without re-paginating the entire vault entry.

Optional dependency (install separately to enable XLSX support):

* ``openpyxl`` — pure-Python XLSX reader.
"""

from __future__ import annotations

import csv
import logging
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any, Protocol, runtime_checkable

from creek.ingest.base import (
    Ingestor,
    ParsedFragment,
    RawDocument,
    file_modified_time,
    parse_authored_at,
)
from creek.ingest.encoding import DEFAULT_CONFIDENCE_THRESHOLD, decode_bytes
from creek.ingest.source_unit import sanitize_unit
from creek.models import SourcePlatform

if TYPE_CHECKING:
    from collections.abc import Sequence

logger = logging.getLogger(__name__)


SPREADSHEET_EXTENSIONS: frozenset[str] = frozenset({".xlsx", ".csv"})
"""File extensions handled by :class:`SpreadsheetIngestor`."""


SUMMARY_THRESHOLD: int = 100
"""Row count above which a sheet renders as a head/tail summary."""

SUMMARY_HEAD_ROWS: int = 10
"""Number of leading rows kept in a summarised table."""

SUMMARY_TAIL_ROWS: int = 5
"""Number of trailing rows kept in a summarised table."""


# ---- Public dataclasses + protocol --------------------------------------


@dataclass(frozen=True)
class SheetData:
    """A single worksheet within a workbook.

    Attributes:
        name: Sheet name as it appears in the workbook.
        headers: Optional header row (auto-detected by the backend);
            ``None`` when the first row is data.
        rows: All data rows as tuples of strings (formulas resolved to
            calculated values, dates ISO-formatted).
    """

    name: str
    headers: tuple[str, ...] | None
    rows: tuple[tuple[str, ...], ...]

    @property
    def is_empty(self) -> bool:
        """``True`` when neither headers nor rows carry any signal."""
        return not self.rows and not self.headers


_BLANK_SHEET_UNIT = "sheet"
"""Stand-in unit for a sheet whose name is blank or whitespace (#1305).

A blank name has nothing to key on, and the two obvious alternatives are
both wrong: an empty discriminator is a second spelling of "the whole
file" (see :func:`~creek.ingest.source_unit.compose_source_unit`), and a
bare positional index is not order-stable, because empty sheets are
skipped and a sheet that gains content shifts every index after it.
Normalising to a name lets the ordinary duplicate-name rule below carry
the blank case too.
"""


def _sheet_unit_keys(sheets: Sequence[SheetData]) -> list[str | None]:
    """Return the per-sheet source-unit discriminator for *sheets* (#1305).

    Two rules, in this order.

    **A single non-empty sheet gets no discriminator at all.** This is the
    migration answer, achieved by construction rather than by a backfill:
    every CSV and every one-sheet XLSX derives byte-identical identity
    before and after #1305, so nothing already in an operator's vault is
    re-minted and duplicated on its next ingest. There is no ledger to pin
    spreadsheets into — ``spreadsheet`` is deliberately absent from
    :data:`creek.ingest.pipeline.LEDGERED_SOURCES`, because per-sheet
    sub-unit identity needs its own idempotency proof before it can be
    ledger-backed (#1363) — so a #1329-style migration is unavailable here,
    and this carve-out is what replaces it.

    **Otherwise the discriminator is the sheet's name**, passed through
    :func:`~creek.ingest.source_unit.sanitize_unit`, with a blank name
    normalised to :data:`_BLANK_SHEET_UNIT` and the *n*-th (n>=2)
    occurrence of an already-seen name suffixed ``~<n>``. Excel forbids
    duplicate sheet names, but a hand-built or stub workbook does not, and
    a bare name would silently restore the very collision this fixes.

    The occurrence ordinal is counted over names, not positions, so
    inserting a *differently* named sheet anywhere in the workbook leaves
    every existing sheet's unit — and therefore its fragment id — untouched.
    That order-stability is why the "empty discriminator for sheet index 0"
    shortcut is rejected: an inserted first sheet would inherit the previous
    first sheet's id, and because ``_write_model`` is first-writer-wins
    (``creek/vault/writer.py``) the newcomer would then be silently dropped
    while the real first sheet duplicated.

    Args:
        sheets: The workbook's non-empty sheets, in workbook order.

    Returns:
        One entry per input sheet, positionally aligned: ``None`` for the
        single-sheet case, else the sheet's unit key.
    """
    if len(sheets) < 2:
        return [None] * len(sheets)
    seen: Counter[str] = Counter()
    units: list[str | None] = []
    for sheet in sheets:
        # Sanitised BEFORE counting, not after: Excel permits ``#`` in a
        # sheet title, and two titles that sanitise to the same string
        # (``Rev#2`` and ``Rev-2``) must be seen here as the collision they
        # are. Sanitising downstream of the counter would mint one unit —
        # and so one id — for both, restoring this issue's own defect.
        name = sanitize_unit(sheet.name.strip()) or _BLANK_SHEET_UNIT
        seen[name] += 1
        occurrence = seen[name]
        units.append(name if occurrence == 1 else f"{name}~{occurrence}")
    return units


def _sheet_label(fragment: ParsedFragment) -> str:
    """Return the operator-visible name of *fragment*'s sheet (#1305).

    The **de-duplicated** ``source_unit``, never the raw
    ``metadata["sheet"]``. Excel forbids duplicate sheet titles but a
    hand-built workbook does not, and two sheets both named ``Data`` carry
    the identical raw name. :func:`_sheet_unit_keys` already resolved that
    collision into ``Data`` and ``Data~2``; deriving the title and the body
    heading from the raw name instead hands the operator two fragments with
    distinct ids, distinct files and identical text — disambiguated in the
    index, indistinguishable everywhere they actually read it.

    Consuming the unit also makes the visible label and the ``…#<unit>``
    source key the same string, so a fragment can be matched by eye to the
    ledger and purge records that name it.

    Args:
        fragment: A parsed spreadsheet fragment.

    Returns:
        The fragment's ``source_unit`` when it has one, else the raw sheet
        name. That fallback is reached only by a source carrying no unit at
        all — a CSV, a one-sheet XLSX — which has no sibling to be confused
        with, and whose pre-#1305 rendering is preserved verbatim so no
        fragment already in a vault is re-titled.
    """
    return fragment.source_unit or str(fragment.metadata.get("sheet", "Sheet"))


@dataclass(frozen=True)
class WorkbookData:
    """A workbook (XLSX or CSV) decomposed into sheets.

    A CSV file always presents as a one-sheet workbook whose sheet is
    named after the file stem.
    """

    sheets: tuple[SheetData, ...]


@runtime_checkable
class SpreadsheetBackend(Protocol):
    """Pluggable spreadsheet backend.

    Implementations must be deterministic and side-effect-free; the
    same input file should always produce the same
    :class:`WorkbookData`.
    """

    def is_available(self) -> bool:
        """Return ``True`` when the backend can perform reads right now."""

    def read_workbook(
        self,
        path: Path,
        *,
        has_header: bool | None = None,
    ) -> WorkbookData:
        """Read *path* and return its decomposed :class:`WorkbookData`.

        ``has_header`` lets callers override the per-sheet header
        auto-detection heuristic (#165):

        * ``None`` (default) — auto-detect: a row qualifies as headers
          only when every cell is a non-empty string.
        * ``True`` — always treat the first row of every sheet as a
          header row, regardless of content.
        * ``False`` — never treat the first row as a header; render
          ``col1..colN`` placeholders instead.
        """


class OpenpyxlUnavailableError(RuntimeError):
    """Raised when an XLSX read is attempted without ``openpyxl`` installed."""


# ---- Default backend ---------------------------------------------------


class OpenpyxlBackend:
    """Spreadsheet backend backed by ``openpyxl`` (XLSX) and stdlib ``csv``.

    Imports of the optional ``openpyxl`` dependency are deferred to
    call time so the rest of the package runs cleanly without it.
    CSV files are read via Python's stdlib :mod:`csv` module and
    therefore always work — ``is_available()`` reflects only the
    XLSX path.
    """

    def is_available(self) -> bool:
        """Return ``True`` when ``openpyxl`` imports cleanly."""
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            return False
        return True

    def read_workbook(
        self,
        path: Path,
        *,
        has_header: bool | None = None,
    ) -> WorkbookData:
        """Read *path* and return a :class:`WorkbookData`.

        Args:
            path: Filesystem path to a ``.xlsx`` or ``.csv`` file.
            has_header: Per-call override for header auto-detection
                (#165). ``None`` keeps the heuristic, ``True`` forces
                the first row to be a header, ``False`` forces it to
                be data.

        Returns:
            A :class:`WorkbookData` with one or more
            :class:`SheetData` entries.

        Raises:
            OpenpyxlUnavailableError: When *path* is an XLSX file and
                ``openpyxl`` is not installed.
        """
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return _read_csv(path, has_header=has_header)
        return self._read_xlsx(path, has_header=has_header)

    @staticmethod
    def _read_xlsx(path: Path, *, has_header: bool | None = None) -> WorkbookData:
        """Read an XLSX file via ``openpyxl``."""
        try:
            import openpyxl
        except ImportError as exc:
            msg = (
                "openpyxl is required for XLSX ingestion. Install it with "
                "`pip install openpyxl`."
            )
            raise OpenpyxlUnavailableError(msg) from exc

        workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
        sheets: list[SheetData] = []
        for name in workbook.sheetnames:
            worksheet = workbook[name]
            raw_rows = [
                tuple(_cell_to_str(cell) for cell in row)
                for row in worksheet.iter_rows(values_only=True)
            ]
            sheets.append(_split_header(name, raw_rows, has_header=has_header))
        workbook.close()
        return WorkbookData(sheets=tuple(sheets))


def _cell_to_str(value: Any) -> str:
    """Coerce a workbook cell value to its display string form."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def _extract_xlsx_authored_at(path: Path) -> datetime | None:
    """Read an XLSX file's core-properties ``created`` (then ``modified``).

    FEAT-031: returns the spreadsheet's source-side authored date or
    ``None`` when neither field is populated. CSV files (which carry
    no core properties) and missing-openpyxl environments both
    return ``None`` — the honest answer when no extractable date
    exists; the filesystem mtime is never substituted here.
    """
    if path.suffix.lower() != ".xlsx":
        return None
    workbook = _open_xlsx(path)
    if workbook is None:
        return None
    try:
        props = workbook.properties
        for candidate in (props.created, props.modified):
            parsed = _safe_parse_authored_at(candidate)
            if parsed is not None:
                return parsed
    finally:
        workbook.close()
    return None


def _open_xlsx(path: Path) -> Any | None:
    """Open *path* via openpyxl read-only, swallowing the wide error set.

    Centralising the optional-import + file-open paths here keeps
    :func:`_extract_xlsx_authored_at` linear and below the project's
    per-function try-block cap.
    """
    try:  # noqa: TRY101  # Separate optional-dep absence from file-open failure.
        import openpyxl
    except ImportError:
        return None
    try:  # noqa: TRY101  # Distinct failure mode from the ImportError branch.
        return openpyxl.load_workbook(path, read_only=True)
    except Exception:  # BadZipFile / InvalidFile / OSError / KeyError
        logger.debug("Could not open XLSX core properties for %s", path)
        return None


def _safe_parse_authored_at(candidate: object) -> datetime | None:
    """Wrapper that swallows :class:`ValueError` from ``parse_authored_at``."""
    if candidate is None:
        return None
    try:
        return parse_authored_at(candidate)
    except ValueError:
        return None


CSV_CHARDET_CONFIDENCE_THRESHOLD: float = DEFAULT_CONFIDENCE_THRESHOLD
"""Minimum ``chardet`` confidence required to trust a *single-byte* guess.

This governs the single-byte path only, and is unchanged at 0.7. A
multi-byte detection is admitted below it by codec class instead —
see :mod:`creek.ingest.encoding` for why lowering this number was the
wrong fix and what replaced it (#1589, #1591).

For orientation, the measured verdicts under chardet 7.6.0 that the
threshold alone cannot separate: a GBK CSV scores 0.38 and a short
Shift-JIS one 0.32, both correct and both under the gate; a Cyrillic
CSV scores 0.45 and a genuine cp1252 file's top guess 0.05, both of
which must *not* be trusted.
"""


def _read_csv(path: Path, *, has_header: bool | None = None) -> WorkbookData:
    """Read a CSV file as a single-sheet workbook.

    Decoding is delegated whole to
    :func:`creek.ingest.encoding.decode_bytes`, which owns the probe
    order and the accept/reject rule. When it reports a degraded
    result — no codec positively identified, so a fallback was used —
    this emits a ``WARNING`` naming the file and the guess that was
    rejected, so the user has a chance to spot mojibake before it
    lands in the vault (BUG-010).

    ``has_header`` is threaded into :func:`_split_header` so callers
    can override per-call header auto-detection (#165).

    Args:
        path: The CSV file to read.
        has_header: Optional override for header auto-detection.

    Returns:
        A single-sheet :class:`WorkbookData`.

    Raises:
        UndecodableBytesError: When the file is binary rather than
            text. It propagates to ``Ingestor._parse_safe``, which
            records the failure and skips the file — the loud outcome
            #1591 asked for. Returning garbage instead would write a
            fragment nothing downstream could tell from real text.
    """
    raw = path.read_bytes()
    decoded = decode_bytes(
        raw,
        confidence_threshold=CSV_CHARDET_CONFIDENCE_THRESHOLD,
    )
    if decoded.degraded:
        logger.warning(
            "CSV %s decoded as %s (chardet best guess: %s @ %.2f); "
            "non-Western content may render as mojibake.",
            path,
            decoded.codec,
            decoded.detected or "unknown",
            decoded.confidence,
        )
    return _csv_text_to_workbook(path, decoded.text, has_header=has_header)


def _csv_text_to_workbook(
    path: Path,
    text: str,
    *,
    has_header: bool | None = None,
) -> WorkbookData:
    """Parse decoded CSV ``text`` into a single-sheet :class:`WorkbookData`."""
    rows = [tuple(row) for row in csv.reader(text.splitlines())]
    return WorkbookData(
        sheets=(_split_header(path.stem, rows, has_header=has_header),),
    )


def _split_header(
    name: str,
    rows: list[tuple[str, ...]],
    *,
    has_header: bool | None = None,
) -> SheetData:
    """Split *rows* into ``(headers, data_rows)``.

    Header-row selection follows ``has_header`` (#165):

    * ``None`` (default) auto-detects via the heuristic — the first
      row qualifies when every cell is a non-empty string.
    * ``True`` always treats the first row as headers, regardless of
      content. An empty sheet (``rows == []``) still yields no
      headers since there is no first row to promote.
    * ``False`` never treats the first row as headers; the rendered
      table will use auto-generated ``colN`` placeholders.
    """
    if not rows:
        return SheetData(name=name, headers=None, rows=())
    first = rows[0]
    if has_header is False:
        return SheetData(name=name, headers=None, rows=tuple(rows))
    if has_header is True:
        return SheetData(name=name, headers=first, rows=tuple(rows[1:]))
    if first and all(cell.strip() for cell in first):
        return SheetData(
            name=name,
            headers=first,
            rows=tuple(rows[1:]),
        )
    return SheetData(name=name, headers=None, rows=tuple(rows))


# ---- SpreadsheetIngestor -----------------------------------------------


class SpreadsheetIngestor(Ingestor):
    """Ingest XLSX and CSV files as one fragment per sheet."""

    def __init__(self, backend: SpreadsheetBackend | None = None) -> None:
        """Initialise with a backend; defaults to :class:`OpenpyxlBackend`."""
        self.backend = backend if backend is not None else OpenpyxlBackend()

    def discover(self, source_path: Path) -> list[RawDocument]:
        """Recursively find ``.xlsx`` and ``.csv`` files under *source_path*.

        Single-file paths whose suffix is unsupported return an empty
        list so the ingest pipeline can route them elsewhere. File
        bytes are not slurped at discovery time — each backend reads
        from disk via the :class:`Path`.
        """
        if source_path.is_file():
            paths = (
                [source_path]
                if source_path.suffix.lower() in SPREADSHEET_EXTENSIONS
                else []
            )
        else:
            paths = [
                p
                for p in sorted(source_path.rglob("*"))
                if p.is_file() and p.suffix.lower() in SPREADSHEET_EXTENSIONS
            ]
        return [
            RawDocument(
                path=path,
                content=b"",
                metadata={"original_file": str(path)},
                detected_encoding="binary"
                if path.suffix.lower() == ".xlsx"
                else "utf-8",
            )
            for path in paths
        ]

    def parse(
        self,
        raw: RawDocument,
        *,
        has_header: bool | None = None,
    ) -> list[ParsedFragment]:
        """Read the workbook and emit one fragment per non-empty sheet.

        ``has_header`` lets callers override per-document header
        auto-detection (#165). The default ``None`` preserves the
        heuristic; ``True`` / ``False`` force the first row to be
        treated as headers / data respectively, in every sheet of the
        workbook.

        FEAT-031: when the workbook is an XLSX, the spreadsheet's core
        properties ``created`` (then ``modified``) become
        ``authored_at`` on every sheet's fragment. CSV files carry no
        authored-date metadata so they fall through to ``None`` — the
        filesystem mtime is *not* a substitute.
        """
        workbook = self.backend.read_workbook(raw.path, has_header=has_header)
        timestamp = file_modified_time(raw.path)
        authored_at = _extract_xlsx_authored_at(raw.path)
        # Materialised before the loop because the *number* of non-empty
        # sheets decides whether a discriminator is emitted at all (#1305).
        sheets = [sheet for sheet in workbook.sheets if not sheet.is_empty]
        units = _sheet_unit_keys(sheets)
        fragments: list[ParsedFragment] = []
        for sheet, unit in zip(sheets, units, strict=True):
            column_count = (
                len(sheet.headers)
                if sheet.headers
                else (len(sheet.rows[0]) if sheet.rows else 0)
            )
            fragments.append(
                ParsedFragment(
                    content="",  # Markdown rendered in convert_to_markdown.
                    metadata={
                        "original_file": str(raw.path),
                        "sheet": sheet.name,
                        "rows": len(sheet.rows),
                        "columns": column_count,
                        "authored_at": authored_at,
                    },
                    # The whole file, always. The per-sheet discriminator
                    # rides on ``source_unit``; overloading this string would
                    # break `creek.ingest.routing.arbitrate`, which decides
                    # ingestor ownership by grouping on it (#1304).
                    source_path=str(raw.path),
                    source_unit=unit,
                    timestamp=timestamp,
                    payload=sheet,
                ),
            )
        return fragments

    def convert_to_markdown(self, fragment: ParsedFragment) -> str:
        """Render the sheet as a GFM table with optional summary truncation.

        #1305: the heading names the sheet by its **de-duplicated unit**
        via :func:`_sheet_label`, not by the raw sheet name. See that
        function for why the distinction is the whole point.
        """
        headers, rows = _extract_headers_and_rows(fragment)
        original = Path(
            str(fragment.metadata.get("original_file", fragment.source_path))
        )
        lines = [f"# {original.name} — {_sheet_label(fragment)}", ""]
        if not headers and not rows:
            lines.append("_(empty sheet)_")
            return "\n".join(lines) + "\n"
        lines.extend(_render_table(headers, rows))
        return "\n".join(lines) + "\n"

    def generate_frontmatter(self, fragment: ParsedFragment) -> dict[str, Any]:
        """Produce YAML frontmatter for a spreadsheet fragment.

        FEAT-031: ``authored_at`` (XLSX core properties ``created`` /
        ``modified``) lands on the frontmatter as an ISO string when
        present; absent for CSV and for XLSX files whose core
        properties carry no creation date.

        #1305: ``title`` is now always emitted, and names the sheet — by
        its de-duplicated unit, see :func:`_sheet_label` — when the
        fragment carries one. Two reasons it cannot be left
        to the ``setdefault`` fallback in
        :func:`~creek.ingest.base.assemble_ingested_fragment`. First, that
        fallback is ``Path(source_path).stem``, which is the *workbook*
        for every sheet — the fallback does not accidentally save us.
        Second, ``sheet`` / ``rows`` / ``columns`` above never reach a
        file: :class:`~creek.models.Fragment` leaves pydantic's default
        ``extra="ignore"``, so ``Fragment.model_validate`` drops them.
        After #1305 the title, the filename it produces, and the body
        heading are the only per-sheet markers in the vault, which makes
        emitting one a correctness requirement rather than a nicety. That
        drop is accepted here and tracked in #1392 rather than fixed
        under a bugfix, because widening what survives validation is a
        repo-wide frontmatter change.

        A fragment with **no** unit — a CSV, a one-sheet XLSX — gets the
        bare stem, which is byte-identical to what the fallback produced
        before this change. Naming the lone sheet there would move the
        computed filename of every such fragment already in a vault; the
        id pin is only half a pin if the filename still churns.
        """
        original_file = str(
            fragment.metadata.get("original_file", fragment.source_path)
        )
        stem = Path(original_file).stem
        title = f"{stem} — {_sheet_label(fragment)}" if fragment.source_unit else stem
        frontmatter_dict: dict[str, Any] = {
            "type": "fragment",
            "title": title,
            "source": {
                "platform": SourcePlatform.SPREADSHEET.value,
                "original_file": original_file,
            },
            "sheet": fragment.metadata.get("sheet", ""),
            "rows": fragment.metadata.get("rows", 0),
            "columns": fragment.metadata.get("columns", 0),
            "ingested": fragment.timestamp.isoformat(),
        }
        authored_at: datetime | None = fragment.metadata.get("authored_at")
        if authored_at is not None:
            frontmatter_dict["authored_at"] = authored_at.isoformat()
        return frontmatter_dict


def _extract_headers_and_rows(
    fragment: ParsedFragment,
) -> tuple[list[str], list[list[str]]]:
    """Return ``(headers, rows)`` from a parsed fragment's typed payload.

    Reads the typed :class:`SheetData` off
    :attr:`~creek.ingest.base.ParsedFragment.payload` to avoid the
    string-keyed metadata round-trip ``"headers"`` / ``"row_data"``
    that the pre-refactor code paid (issue #166).
    """
    sheet = fragment.payload
    if not isinstance(sheet, SheetData):
        return [], []
    headers: list[str] = list(sheet.headers) if sheet.headers else []
    rows: list[list[str]] = [list(row) for row in sheet.rows]
    if not headers and rows:
        headers = [f"col{i + 1}" for i in range(len(rows[0]))]
    return headers, rows


def _escape_cell(value: object) -> str:
    """Escape a cell value for safe inclusion in a GFM table.

    GFM uses ``|`` as the column separator and a literal newline as
    the row terminator, so cells carrying either character would
    corrupt the table layout. We escape ``|`` to ``\\|`` (the GFM
    canonical form) and replace any newline with ``<br>`` so the
    cell renders on a single visual line.
    """
    return (
        str(value)
        .replace("|", r"\|")
        .replace("\r\n", "<br>")
        .replace("\n", "<br>")
        .replace("\r", "<br>")
    )


def _render_row(cells: list[str] | tuple[str, ...]) -> str:
    """Render one GFM table row from a sequence of pre-escaped cells."""
    return "| " + " | ".join(cells) + " |"


def _render_table(headers: list[str], rows: list[list[str]]) -> list[str]:
    """Return markdown lines for a GFM table, summarising oversize sheets.

    Cells are escaped via :func:`_escape_cell` so values containing
    ``|`` or newlines do not corrupt the table. Sheets larger than
    :data:`SUMMARY_THRESHOLD` are emitted as a single table whose
    body is the head + tail rows; the summary note appears *before*
    the table so the reader sees the truncation context first.
    """
    safe_headers = [_escape_cell(header) for header in headers]
    separator = _render_row(["---"] * len(safe_headers))
    lines: list[str] = []
    if len(rows) > SUMMARY_THRESHOLD:
        lines.extend(
            (
                f"_Showing first {SUMMARY_HEAD_ROWS} and last "
                f"{SUMMARY_TAIL_ROWS} of {len(rows)} rows._",
                "",
            ),
        )
        displayed = rows[:SUMMARY_HEAD_ROWS] + rows[-SUMMARY_TAIL_ROWS:]
    else:
        displayed = rows
    lines.extend((_render_row(safe_headers), separator))
    lines.extend(_render_row([_escape_cell(cell) for cell in row]) for row in displayed)
    return lines


__all__ = [
    "SPREADSHEET_EXTENSIONS",
    "SUMMARY_HEAD_ROWS",
    "SUMMARY_TAIL_ROWS",
    "SUMMARY_THRESHOLD",
    "OpenpyxlBackend",
    "OpenpyxlUnavailableError",
    "SheetData",
    "SpreadsheetBackend",
    "SpreadsheetIngestor",
    "WorkbookData",
]
