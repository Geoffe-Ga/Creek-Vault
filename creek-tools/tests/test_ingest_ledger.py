"""Tests for the per-source ingest ledger + ``source_key`` skeleton (#672).

The ledger maps a stable ``source_key`` (the vault-relative path of a mutable
source unit) to the fragment it produced, the content hash at last ingest, and
a timestamp. This issue is a **no-op pass-through**: it records entries and
populates ``source.origin_key`` but changes no write/skip behaviour — a changed
journal entry still mints a new id here (update-in-place is #673).
"""

from __future__ import annotations

import hashlib
import json
from typing import TYPE_CHECKING

import frontmatter
import pytest

from creek.cli import _run_ingest
from creek.ingest import INGESTOR_REGISTRY
from creek.ingest.ledger import LedgerRecord, SourceLedger
from creek.ingest.pipeline import run_ingest
from creek.ingest.spreadsheets import SpreadsheetIngestor

if TYPE_CHECKING:
    from pathlib import Path

    from creek.ingest.pipeline import IngestRunResult


def _make_vault(tmp_path: Path) -> Path:
    """Scaffold a minimal vault the writer + ledger can target."""
    vault = tmp_path / "vault"
    for d in (
        "00-Creek-Meta/Processing-Log",
        "01-Fragments/Journal",
        "01-Fragments/Notes",
        "personal/journal",
    ):
        (vault / d).mkdir(parents=True, exist_ok=True)
    return vault


def _ingest_markdown(vault: Path, entry: Path) -> tuple[int, list[str], int]:
    """Run the markdown ingestor through the real ``_run_ingest`` seam."""
    return _run_ingest(
        ingestor_cls=INGESTOR_REGISTRY["markdown"],
        source_type="markdown",
        input_path=entry,
        vault_path=vault,
    )


def _written_fragments(vault: Path) -> list[frontmatter.Post]:
    """Load every fragment written under ``01-Fragments``."""
    return [
        frontmatter.load(str(p)) for p in sorted((vault / "01-Fragments").rglob("*.md"))
    ]


# ---- SourceLedger (unit) -----------------------------------------------


class TestSourceLedger:
    """The ledger round-trips records and resolves last-write-wins."""

    def test_path_for_lands_under_meta_state(self, tmp_path: Path) -> None:
        """The ledger file lives under 00-Creek-Meta/State/ingest/."""
        path = SourceLedger.path_for(tmp_path, "markdown")
        assert (
            path == tmp_path / "00-Creek-Meta" / "State" / "ingest" / "markdown.jsonl"
        )

    def test_content_hash_is_deterministic(self) -> None:
        """content_hash is a stable SHA-256 of the content."""
        assert SourceLedger.content_hash("abc") == hashlib.sha256(b"abc").hexdigest()
        assert SourceLedger.content_hash("abc") != SourceLedger.content_hash("abd")

    def test_load_missing_file_is_empty(self, tmp_path: Path) -> None:
        """Loading a ledger with no backing file yields no records."""
        ledger = SourceLedger.load(tmp_path, source="markdown")
        assert ledger.get("anything") is None
        assert len(ledger) == 0

    def test_record_then_reload_roundtrips(self, tmp_path: Path) -> None:
        """A recorded entry survives a reload from disk."""
        ledger = SourceLedger.load(tmp_path, source="markdown")
        ledger.record(
            "a/b.md", "frag-aaa", "hash1", last_seen="2026-06-26T00:00:00+00:00"
        )

        reloaded = SourceLedger.load(tmp_path, source="markdown")
        rec = reloaded.get("a/b.md")
        assert rec == LedgerRecord(
            source_key="a/b.md",
            fragment_id="frag-aaa",
            content_hash="hash1",
            last_seen="2026-06-26T00:00:00+00:00",
        )

    def test_record_upsert_last_write_wins(self, tmp_path: Path) -> None:
        """Re-recording the same source_key resolves to the latest entry."""
        ledger = SourceLedger.load(tmp_path, source="markdown")
        ledger.record("a/b.md", "frag-aaa", "hash1")
        ledger.record("a/b.md", "frag-aaa", "hash2")

        reloaded = SourceLedger.load(tmp_path, source="markdown")
        rec = reloaded.get("a/b.md")
        assert rec is not None
        assert rec.content_hash == "hash2"

    def test_record_auto_stamps_last_seen(self, tmp_path: Path) -> None:
        """Omitting last_seen stamps an ISO-8601 timestamp."""
        ledger = SourceLedger.load(tmp_path, source="markdown")
        rec = ledger.record("a/b.md", "frag-aaa", "hash1")
        assert rec.last_seen  # non-empty ISO timestamp
        assert "T" in rec.last_seen

    def test_malformed_lines_are_skipped(self, tmp_path: Path) -> None:
        """A corrupt JSONL line does not abort loading the rest."""
        path = SourceLedger.path_for(tmp_path, "markdown")
        path.parent.mkdir(parents=True, exist_ok=True)
        good = json.dumps(
            {
                "source_key": "a/b.md",
                "fragment_id": "frag-aaa",
                "content_hash": "h",
                "last_seen": "t",
            }
        )
        path.write_text(f"{{not json\n\n{good}\n", encoding="utf-8")
        ledger = SourceLedger.load(tmp_path, source="markdown")
        assert ledger.get("a/b.md") is not None

    def test_records_missing_required_fields_are_rejected(self, tmp_path: Path) -> None:
        """A row missing a required field (e.g. content_hash) is not loaded."""
        path = SourceLedger.path_for(tmp_path, "markdown")
        path.parent.mkdir(parents=True, exist_ok=True)
        partial = json.dumps(
            {"source_key": "a/b.md", "fragment_id": "frag-aaa", "last_seen": "t"}
        )
        empty_hash = json.dumps(
            {
                "source_key": "c/d.md",
                "fragment_id": "frag-bbb",
                "content_hash": "",
                "last_seen": "t",
            }
        )
        path.write_text(f"{partial}\n{empty_hash}\n", encoding="utf-8")
        ledger = SourceLedger.load(tmp_path, source="markdown")
        assert ledger.get("a/b.md") is None  # missing content_hash
        assert ledger.get("c/d.md") is None  # empty content_hash
        assert len(ledger) == 0


# ---- Markdown ingest wiring (no-op pass-through) -----------------------


class TestMarkdownLedgerWiring:
    """Markdown ingest populates origin_key + the ledger, behaviour unchanged."""

    def test_origin_key_and_ledger_recorded(self, tmp_path: Path) -> None:
        """An ingested journal entry carries origin_key and a ledger record."""
        vault = _make_vault(tmp_path)
        entry = vault / "personal" / "journal" / "2026-06-26.md"
        entry.write_text(
            "---\ndate: 2026-06-26\n---\nFirst draft of today.\n",
            encoding="utf-8",
        )

        written, errors, _ = _ingest_markdown(vault, entry)
        assert written == 1, errors

        frags = _written_fragments(vault)
        assert len(frags) == 1
        post = frags[0]
        assert post["source"]["origin_key"] == "personal/journal/2026-06-26.md"

        ledger = SourceLedger.load(vault, source="markdown")
        rec = ledger.get("personal/journal/2026-06-26.md")
        assert rec is not None
        assert rec.fragment_id == post["id"]
        assert rec.content_hash  # non-empty
        assert rec.last_seen

    def test_reingest_unchanged_is_still_a_noop(self, tmp_path: Path) -> None:
        """Re-ingesting unchanged input writes no duplicate (skeleton invariant)."""
        vault = _make_vault(tmp_path)
        entry = vault / "personal" / "journal" / "2026-06-26.md"
        entry.write_text(
            "---\ndate: 2026-06-26\n---\nStable content.\n",
            encoding="utf-8",
        )

        _ingest_markdown(vault, entry)
        _ingest_markdown(vault, entry)

        assert len(_written_fragments(vault)) == 1

    def test_non_markdown_source_skips_ledger(self, tmp_path: Path) -> None:
        """Only the markdown source writes a ledger in this skeleton."""
        vault = _make_vault(tmp_path)
        # No markdown ingest run -> no ledger file should exist.
        assert not SourceLedger.path_for(vault, "discord").exists()


# ---- Per-sheet ledger identity for uploaded workbooks (#1305) ----------


_UPLOAD_RELDIR = "00-Creek-Meta/adepthood/uploads"
"""Vault-relative staging directory the upload path writes into."""

_EXPECTED_SHEET_KEYS = {
    f"{_UPLOAD_RELDIR}/book.xlsx#Budget",
    f"{_UPLOAD_RELDIR}/book.xlsx#Notes",
    f"{_UPLOAD_RELDIR}/book.xlsx#Q3",
}
"""The three per-sheet ``source_key`` values one staged workbook must produce."""


def _stage_multi_sheet_workbook(vault: Path) -> Path:
    """Write a real 3-sheet workbook into the upload staging directory.

    Each sheet carries a header row and a data row so none of them is
    dropped by the ``sheet.is_empty`` filter in ``SpreadsheetIngestor.parse``.
    """
    import openpyxl

    staged = vault / _UPLOAD_RELDIR / "book.xlsx"
    staged.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    first = workbook.active
    assert first is not None
    first.title = "Budget"
    workbook.create_sheet(title="Notes")
    workbook.create_sheet(title="Q3")
    for index, name in enumerate(("Budget", "Notes", "Q3")):
        worksheet = workbook[name]
        worksheet.append(["item", "amount"])
        worksheet.append([f"row-{index}", str(index * 10)])
    workbook.save(staged)
    return staged


def _ingest_staged_workbook(vault: Path, staged: Path) -> IngestRunResult:
    """Run the spreadsheet ingestor over the staged upload, ledger-backed."""
    return run_ingest(
        ingestor_cls=SpreadsheetIngestor,
        source_type="spreadsheet",
        input_path=staged,
        vault_path=vault,
        ledger_source="upload",
    )


def _fragment_paths(vault: Path) -> list[Path]:
    """Return every fragment markdown file under ``01-Fragments``, sorted."""
    return sorted((vault / "01-Fragments").rglob("*.md"))


class TestUploadedWorkbookLedgerIdentity:
    """A staged multi-sheet workbook is one ledgered unit per sheet (#1305)."""

    def test_ledgered_multi_sheet_workbook_is_created_then_unchanged(
        self,
        tmp_path: Path,
    ) -> None:
        """Three sheets create three fragments, then re-ingest as three unchanged.

        Measured at HEAD, this run reports::

            run1  created=1 updated=0 unchanged=2  files=1
            run2  created=0 updated=0 unchanged=3  files=1

        and the one surviving file holds the FIRST sheet (``Budget``) —
        ``VaultWriter._write_model``'s duplicate-id early return
        returns the already-written path for a duplicate id, so the first
        writer wins and sheets 2..N are dropped. Any assertion phrased
        around "the last sheet survives" would be red before *and* after
        the fix and would certify nothing.

        A per-sheet *fragment id* alone does not fix this run, which is
        why the ledgered path needs its own test. ``derive_source_key``
        keys on the file, so all three sheets resolve to one
        ``origin_key``, one ``LedgerRecord``, and therefore one identity:
        ``write_fragment_idempotent`` assigns ``fragment.id =
        record.fragment_id`` on every branch of ``write_fragment_idempotent``
        and the freshly-derived per-sheet ids are overwritten before the
        write. The ledger key has to become per-sheet too.

        ``updated`` must stay at zero on both runs. Nothing was edited
        between them, so an ``updated`` count here is three sheets
        overwriting each other in place — the loss wearing a success
        message.
        """
        vault = _make_vault(tmp_path)
        staged = _stage_multi_sheet_workbook(vault)

        run1 = _ingest_staged_workbook(vault, staged)
        assert run1.errors == [], run1.errors
        assert (run1.created, run1.updated, run1.unchanged) == (3, 0, 0)
        assert len(_fragment_paths(vault)) == 3

        run2 = _ingest_staged_workbook(vault, staged)
        assert run2.errors == [], run2.errors
        assert (run2.created, run2.updated, run2.unchanged) == (0, 0, 3)
        assert len(_fragment_paths(vault)) == 3

    def test_ledgered_sheet_origin_keys_are_per_sheet_and_stable(
        self,
        tmp_path: Path,
    ) -> None:
        """Each sheet gets its own ``origin_key``, and it does not move on re-ingest.

        ``source.origin_key`` is the key the RTBF purge sweep matches on,
        so it is not merely an identity nicety: a workbook whose three
        sheets share one key is three fragments the operator can only
        delete as one, and a key that changes between two identical
        ingests is a fragment the purge can no longer find at all.

        At HEAD all three sheets derive the same key
        (``00-Creek-Meta/adepthood/uploads/book.xlsx``), so the ledger
        holds exactly one record and this test is red on the key set.
        """
        vault = _make_vault(tmp_path)
        staged = _stage_multi_sheet_workbook(vault)

        first = _ingest_staged_workbook(vault, staged)
        assert first.errors == [], first.errors

        ledger = SourceLedger.load(vault, source="upload")
        assert len(ledger) == 3
        assert ledger.live_keys() == _EXPECTED_SHEET_KEYS

        posts = _written_fragments(vault)
        assert len(posts) == 3
        assert {post["source"]["origin_key"] for post in posts} == _EXPECTED_SHEET_KEYS
        for post in posts:
            record = ledger.get(post["source"]["origin_key"])
            assert record is not None, post["source"]["origin_key"]
            assert record.fragment_id == post["id"]

        second = _ingest_staged_workbook(vault, staged)
        assert second.errors == [], second.errors
        reloaded = SourceLedger.load(vault, source="upload")
        assert reloaded.live_keys() == _EXPECTED_SHEET_KEYS


# ---------------------------------------------------------------------------
# forget_fragment_ids — the erasure path (#1453)
# ---------------------------------------------------------------------------

_FORGET_HASH = "e29568dd0772c2a8ac12fc4a677d3b6a678baa35f8b17fe0f7c76b59dacd3335"
"""A full unsalted SHA-256, the shape the ledger really stores."""


def _write_rows(vault: Path, source: str, rows: list[dict[str, object]]) -> Path:
    """Hand-write raw ledger rows, bypassing :class:`SourceLedger`.

    Args:
        vault: Vault root.
        source: Ingestor key naming the ``<source>.jsonl`` file.
        rows: Row objects, written one per line in order.

    Returns:
        Path to the ledger file.
    """
    path = SourceLedger.path_for(vault, source)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "".join(f"{json.dumps(row)}\n" for row in rows),
        encoding="utf-8",
    )
    return path


def _row(source_key: str, fragment_id: str, content_hash: str) -> dict[str, object]:
    """Build one well-formed ledger row.

    Args:
        source_key: Stable vault-relative identity of the source unit.
        fragment_id: The fragment id it maps to.
        content_hash: SHA-256 of the source content.

    Returns:
        A row object ready for :func:`_write_rows`.
    """
    return {
        "source_key": source_key,
        "fragment_id": fragment_id,
        "content_hash": content_hash,
        "last_seen": "2026-08-13T00:00:00+00:00",
        "tombed": False,
    }


def test_forget_removes_the_row_naming_the_doomed_id(tmp_path: Path) -> None:
    """The row mapping source path to purged id is physically gone (#1453)."""
    from creek.ingest.ledger import forget_fragment_ids

    vault = _make_vault(tmp_path)
    ledger = _write_rows(
        vault,
        "markdown",
        [
            _row("therapy.md", "frag-doomed", _FORGET_HASH),
            _row("recipes.md", "frag-safe", "b" * 64),
        ],
    )

    removed = forget_fragment_ids(vault, ["frag-doomed"])

    assert removed == 1
    text = ledger.read_text(encoding="utf-8")
    assert "frag-doomed" not in text
    assert "therapy.md" not in text
    assert _FORGET_HASH not in text
    assert "frag-safe" in text


def test_forget_takes_the_whole_history_of_a_doomed_source_unit(
    tmp_path: Path,
) -> None:
    """Superseded rows for the same unit go too (#1453).

    The ledger is append-only, so one source unit accumulates a row per
    ingest and each carries that revision's hash. Matching on
    ``fragment_id`` alone leaves hashes of *earlier drafts of the same
    private text* on disk, still joined to the source path.
    """
    from creek.ingest.ledger import forget_fragment_ids

    vault = _make_vault(tmp_path)
    ledger = _write_rows(
        vault,
        "markdown",
        [
            _row("therapy.md", "frag-v1", "1" * 64),
            _row("therapy.md", "frag-v2", "2" * 64),
            _row("therapy.md", "frag-current", _FORGET_HASH),
            _row("recipes.md", "frag-safe", "b" * 64),
        ],
    )

    removed = forget_fragment_ids(vault, ["frag-current"])

    assert removed == 3
    text = ledger.read_text(encoding="utf-8")
    assert "therapy.md" not in text
    for digest in ("1" * 64, "2" * 64, _FORGET_HASH):
        assert digest not in text
    assert text == f"{json.dumps(_row('recipes.md', 'frag-safe', 'b' * 64))}\n"


def test_forget_removes_an_unparseable_line_that_still_names_the_id(
    tmp_path: Path,
) -> None:
    """A half-written row is the case a structured matcher cannot see (#1453).

    A crash mid-append leaves a truncated line. It is not valid JSON, so
    every reader skips it — and it still spells the fragment id in
    cleartext, which is exactly what an erasure has to remove.
    """
    from creek.ingest.ledger import forget_fragment_ids

    vault = _make_vault(tmp_path)
    path = SourceLedger.path_for(vault, "markdown")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        f'{{"source_key": "therapy.md", "fragment_id": "frag-doomed", "conte\n'
        f"{json.dumps(_row('recipes.md', 'frag-safe', 'b' * 64))}\n",
        encoding="utf-8",
    )

    removed = forget_fragment_ids(vault, ["frag-doomed"])

    assert removed == 1
    assert "frag-doomed" not in path.read_text(encoding="utf-8")


def test_forget_unlinks_a_ledger_left_with_no_rows(tmp_path: Path) -> None:
    """An emptied ledger file is removed, not left as a husk naming its source."""
    from creek.ingest.ledger import forget_fragment_ids

    vault = _make_vault(tmp_path)
    ledger = _write_rows(
        vault,
        "markdown",
        [_row("therapy.md", "frag-doomed", _FORGET_HASH)],
    )

    assert forget_fragment_ids(vault, ["frag-doomed"]) == 1
    assert not ledger.exists()


def test_forget_dry_run_counts_and_leaves_every_byte(tmp_path: Path) -> None:
    """The preview predicts its apply twin exactly and writes nothing."""
    from creek.ingest.ledger import forget_fragment_ids

    vault = _make_vault(tmp_path)
    ledger = _write_rows(
        vault,
        "markdown",
        [
            _row("therapy.md", "frag-v1", "1" * 64),
            _row("therapy.md", "frag-doomed", _FORGET_HASH),
            _row("recipes.md", "frag-safe", "b" * 64),
        ],
    )
    before = ledger.read_bytes()

    dry = forget_fragment_ids(vault, ["frag-doomed"], dry_run=True)

    assert ledger.read_bytes() == before
    assert dry == forget_fragment_ids(vault, ["frag-doomed"])


def test_forget_sweeps_every_source_ledger_not_just_one(tmp_path: Path) -> None:
    """A fragment can be named by more than one ``<source>.jsonl``."""
    from creek.ingest.ledger import forget_fragment_ids

    vault = _make_vault(tmp_path)
    md = _write_rows(vault, "markdown", [_row("a.md", "frag-doomed", "1" * 64)])
    up = _write_rows(vault, "upload", [_row("b.docx", "frag-doomed", "2" * 64)])

    assert forget_fragment_ids(vault, ["frag-doomed"]) == 2
    assert not md.exists()
    assert not up.exists()


def test_forget_with_no_ids_is_a_no_op_never_a_wildcard(tmp_path: Path) -> None:
    """An empty id collection erases nothing.

    The dangerous reading of "forget these ids" with an empty list is
    "forget everything". A scoped purge that matched no fragments hands
    exactly that, on every run.
    """
    from creek.ingest.ledger import forget_fragment_ids

    vault = _make_vault(tmp_path)
    ledger = _write_rows(vault, "markdown", [_row("a.md", "frag-safe", "1" * 64)])
    before = ledger.read_bytes()

    assert forget_fragment_ids(vault, []) == 0
    assert forget_fragment_ids(vault, ["", "   "]) == 0
    assert ledger.read_bytes() == before


def test_forget_tolerates_a_vault_with_no_ledger_directory(tmp_path: Path) -> None:
    """A vault that has never ingested is not an error condition."""
    from creek.ingest.ledger import forget_fragment_ids

    assert forget_fragment_ids(_make_vault(tmp_path), ["frag-doomed"]) == 0


def test_forget_leaves_no_temp_file_behind(tmp_path: Path) -> None:
    """The atomic rewrite cleans up after itself.

    The temp file holds the surviving rows *and* is written into the
    ledger directory, so one left behind is both litter and a second
    copy of data the next erasure would not know to look for.
    """
    from creek.ingest.ledger import forget_fragment_ids, ledger_dir

    vault = _make_vault(tmp_path)
    _write_rows(
        vault,
        "markdown",
        [
            _row("therapy.md", "frag-doomed", _FORGET_HASH),
            _row("recipes.md", "frag-safe", "b" * 64),
        ],
    )

    forget_fragment_ids(vault, ["frag-doomed"])

    assert sorted(p.name for p in ledger_dir(vault).iterdir()) == ["markdown.jsonl"]


def test_forget_drops_blank_lines_without_counting_them(tmp_path: Path) -> None:
    """A blank line is not a row, so it is neither erased nor counted.

    It still must not survive the rewrite: a file whose only remaining
    content is whitespace would keep the ledger *file* alive, and the
    filename itself names the source type.
    """
    from creek.ingest.ledger import forget_fragment_ids

    vault = _make_vault(tmp_path)
    path = SourceLedger.path_for(vault, "markdown")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        f"\n{json.dumps(_row('therapy.md', 'frag-doomed', _FORGET_HASH))}\n\n",
        encoding="utf-8",
    )

    assert forget_fragment_ids(vault, ["frag-doomed"]) == 1
    assert not path.exists()


def test_a_failed_rewrite_leaves_no_temp_file_and_re_raises(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """A crash mid-rewrite cleans up its temp file and does not swallow (#1453).

    The temp file holds the *surviving* rows, so one abandoned in the
    ledger directory is a second copy of data the next erasure has no
    reason to look at. And the failure must propagate: ``_run_audited``
    turns it into a ``status="partial"`` outcome line, and an erasure
    that silently failed while reporting success is the worst outcome
    available here.
    """
    from creek.ingest import ledger as ledger_module
    from creek.ingest.ledger import forget_fragment_ids, ledger_dir

    vault = _make_vault(tmp_path)
    _write_rows(
        vault,
        "markdown",
        [
            _row("therapy.md", "frag-doomed", _FORGET_HASH),
            _row("recipes.md", "frag-safe", "b" * 64),
        ],
    )

    def _boom(self: Path, target: Path) -> Path:
        """Fail the atomic replace.

        Args:
            self: The temp file.
            target: The ledger path.

        Raises:
            OSError: Always.
        """
        msg = "disk went away"
        raise OSError(msg)

    monkeypatch.setattr(ledger_module.Path, "replace", _boom)

    with pytest.raises(OSError, match="disk went away"):
        forget_fragment_ids(vault, ["frag-doomed"])

    assert sorted(p.name for p in ledger_dir(vault).iterdir()) == ["markdown.jsonl"]
